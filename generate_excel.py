# generate_excel.py
# Naitik Singhal, IIT Kanpur

import sys, os
sys.path.append(os.path.join(os.path.dirname(__file__), '..', 'data'))

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
from sklearn.preprocessing import StandardScaler
from sklearn.cluster import KMeans
from indian_chemicals_data import build_master_df

OUT = "/home/claude/indian-chemical-sector-analysis/reports/Indian_Chemical_Sector_Analysis.xlsx"
os.makedirs(os.path.dirname(OUT), exist_ok=True)

df = build_master_df()

# clustering
feats = ["Rev_CAGR_5Y","Avg_EBITDA","Avg_PAT","Avg_ROE",
         "Export_Pct","RnD_Pct","Capex_Pct"]
Xsc = StandardScaler().fit_transform(df[feats].values)
km  = KMeans(n_clusters=4, random_state=42, n_init=10)
km.fit(Xsc)
df["Cluster"] = km.labels_
order = df.groupby("Cluster")["Avg_EBITDA"].mean().sort_values(ascending=False).index
remap = {o:n+1 for n,o in enumerate(order)}
df["Cluster"] = df["Cluster"].map(remap)
cnames = {1:"Premium Specialty",2:"Agro-Export Champions",
          3:"Transitioning Mid-tier",4:"Commodity Players"}
df["Cluster_Name"] = df["Cluster"].map(cnames)

# composite score
for c in ["Avg_EBITDA","Avg_ROE","Rev_CAGR_5Y","RnD_Pct"]:
    mn,mx = df[c].min(), df[c].max()
    df[c+"_n"] = (df[c]-mn)/(mx-mn)*100
df["Score"] = (0.40*df["Avg_EBITDA_n"] + 0.30*df["Avg_ROE_n"] +
               0.20*df["Rev_CAGR_5Y_n"] + 0.10*df["RnD_Pct_n"])

YEARS = [2019,2020,2021,2022,2023,2024]
ECOLS = [f"EBITDA_FY{str(y)[2:]}" for y in YEARS]
RCOLS = [f"Rev_FY{str(y)[2:]}"    for y in YEARS]
ROECOLS=[f"ROE_FY{str(y)[2:]}"    for y in YEARS]

# style helpers - keeping it minimal
def thin():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def write_cell(ws, row, col, value, bold=False, sz=10,
               align="center", wrap=False,
               bg=None, fg="000000", num_fmt=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=bold, size=sz, color=fg)
    c.alignment = Alignment(horizontal=align, vertical="center",
                            wrap_text=wrap)
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    if num_fmt:
        c.number_format = num_fmt
    c.border = thin()
    return c

def col_w(ws, widths):
    for i,w in enumerate(widths,1):
        ws.column_dimensions[get_column_letter(i)].width = w

wb = Workbook()
wb.remove(wb.active)

# ---------------------------------------------------------------
# SHEET 1: Master Data
# ---------------------------------------------------------------
ws = wb.create_sheet("Master Data")
ws.sheet_view.showGridLines = False

# title row
ws.merge_cells("A1:S1")
c = ws["A1"]
c.value = "Indian Chemical Sector - Financial Data (FY2019 to FY2024) | 22 NSE-listed Companies"
c.font = Font(bold=True, size=12, color="FFFFFF")
c.fill = PatternFill("solid", fgColor="2C3E50")
c.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 26

# source row
ws.merge_cells("A2:S2")
ws["A2"].value = "Sources: Annual Reports, Screener.in, NSE disclosures. Figures in INR Crores unless stated."
ws["A2"].font = Font(size=8, italic=True, color="666666")
ws["A2"].alignment = Alignment(horizontal="center")
ws.row_dimensions[2].height = 15

hdrs = ["#","Ticker","Company","Sub-Sector",
        "Rev FY19","Rev FY20","Rev FY21","Rev FY22","Rev FY23","Rev FY24",
        "Rev CAGR 5Y (%)","Avg EBITDA (%)","Avg PAT (%)","Avg ROE (%)",
        "Export (%)","R&D (%)","Capex (%)","Score","Cluster Name"]

for j,h in enumerate(hdrs,1):
    write_cell(ws, 3, j, h, bold=True, sz=9, bg="ECF0F1", wrap=True)
ws.row_dimensions[3].height = 30

dfs = df.sort_values("Score", ascending=False).reset_index(drop=True)
for i, row in dfs.iterrows():
    r = i+4
    bg = "F8F9F9" if i%2==0 else "FFFFFF"
    data = [i+1, row["Ticker"], row["Company"], row["Sub_Sector"],
            int(row["Rev_FY19"]), int(row["Rev_FY20"]), int(row["Rev_FY21"]),
            int(row["Rev_FY22"]), int(row["Rev_FY23"]), int(row["Rev_FY24"]),
            round(row["Rev_CAGR_5Y"],1),
            round(row["Avg_EBITDA"],1), round(row["Avg_PAT"],1),
            round(row["Avg_ROE"],1), row["Export_Pct"],
            row["RnD_Pct"], row["Capex_Pct"],
            round(row["Score"],1), row["Cluster_Name"]]
    aligns = ["center","center","left","left"] + ["center"]*15
    for j,(v,al) in enumerate(zip(data,aligns),1):
        nf = "#,##0" if j in [5,6,7,8,9,10] else None
        write_cell(ws, r, j, v, align=al, bg=bg, num_fmt=nf)

# conditional formatting on EBITDA column (L)
ws.conditional_formatting.add(
    f"L4:L{3+len(df)}",
    ColorScaleRule(start_type="min", start_color="FFFFFF",
                   end_type="max",   end_color="27AE60"))
ws.freeze_panes = "E4"
col_w(ws, [3,9,22,16,9,9,9,9,9,9,10,11,10,9,9,8,8,8,22])

# ---------------------------------------------------------------
# SHEET 2: Sub-sector Summary
# ---------------------------------------------------------------
ws2 = wb.create_sheet("Sub-sector Summary")
ws2.sheet_view.showGridLines = False

ws2.merge_cells("A1:K1")
c = ws2["A1"]
c.value = "Sub-sector Aggregate Analysis (FY2019 to FY2024)"
c.font = Font(bold=True, size=12, color="FFFFFF")
c.fill = PatternFill("solid", fgColor="1A5276")
c.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 26

agg = df.groupby("Sub_Sector").agg(
    N=("Company","count"),
    RevFY19=("Rev_FY19","sum"), RevFY24=("Rev_FY24","sum"),
    EBITDA=("Avg_EBITDA","mean"), PAT=("Avg_PAT","mean"),
    ROE=("Avg_ROE","mean"), Export=("Export_Pct","mean"),
    RnD=("RnD_Pct","mean"), Capex=("Capex_Pct","mean"),
    RevCAGR=("Rev_CAGR_5Y","mean")
).reset_index()
agg["SectorCAGR"] = ((agg["RevFY24"]/agg["RevFY19"])**(1/5)-1)*100

h2 = ["Sub-Sector","No. of Cos","Agg Rev FY19\n(INR Cr)",
      "Agg Rev FY24\n(INR Cr)","Sector CAGR (%)","Avg EBITDA (%)","Avg PAT (%)",
      "Avg ROE (%)","Avg Export (%)","Avg R&D (%)","Avg Capex (%)"]
for j,h in enumerate(h2,1):
    write_cell(ws2, 2, j, h, bold=True, sz=9, bg="ECF0F1", wrap=True)
ws2.row_dimensions[2].height = 36

for i, row in agg.iterrows():
    r = i+3
    vals = [row["Sub_Sector"], int(row["N"]),
            int(row["RevFY19"]), int(row["RevFY24"]),
            round(row["SectorCAGR"],1), round(row["EBITDA"],1),
            round(row["PAT"],1), round(row["ROE"],1),
            round(row["Export"],1), round(row["RnD"],1),
            round(row["Capex"],1)]
    for j,v in enumerate(vals,1):
        nf = "#,##0" if j in [3,4] else None
        al = "left" if j==1 else "center"
        write_cell(ws2, r, j, v, align=al, bg="F8F9F9" if i%2==0 else "FFFFFF", num_fmt=nf)

# key insight row
ws2.row_dimensions[3+len(agg)+1].height = 8
r_note = 3+len(agg)+2
ws2.merge_cells(f"A{r_note}:K{r_note}")
specialty_ebitda = agg.loc[agg["Sub_Sector"]=="Specialty","EBITDA"].values[0]
commodity_ebitda = agg.loc[agg["Sub_Sector"]=="Commodity","EBITDA"].values[0]
ws2[f"A{r_note}"].value = (f"Note: Specialty vs Commodity EBITDA spread = "
    f"{specialty_ebitda-commodity_ebitda:.1f} ppt over FY2019-2024 sample period")
ws2[f"A{r_note}"].font = Font(size=9, italic=True, color="444444")
ws2[f"A{r_note}"].alignment = Alignment(horizontal="left")

col_w(ws2, [18,10,14,14,13,12,11,11,12,10,11])

# ---------------------------------------------------------------
# SHEET 3: EBITDA Trends
# ---------------------------------------------------------------
ws3 = wb.create_sheet("EBITDA Trends")
ws3.sheet_view.showGridLines = False

ws3.merge_cells("A1:J1")
ws3["A1"].value = "EBITDA Margin by Company and Year (%)"
ws3["A1"].font = Font(bold=True, size=12, color="FFFFFF")
ws3["A1"].fill = PatternFill("solid", fgColor="117A65")
ws3["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws3.row_dimensions[1].height = 26

h3 = ["Ticker","Company","Sub-Sector",
      "FY19","FY20","FY21","FY22","FY23","FY24","Change FY19 to FY24"]
for j,h in enumerate(h3,1):
    write_cell(ws3, 2, j, h, bold=True, sz=9, bg="ECF0F1", wrap=True)
ws3.row_dimensions[2].height = 30

dfe = df.sort_values("Avg_EBITDA", ascending=False).reset_index(drop=True)
for i, row in dfe.iterrows():
    r = i+3
    bg = "F8F9F9" if i%2==0 else "FFFFFF"
    write_cell(ws3, r, 1, row["Ticker"], bg=bg)
    write_cell(ws3, r, 2, row["Company"], align="left", bg=bg)
    write_cell(ws3, r, 3, row["Sub_Sector"], align="left", bg=bg)
    for k,ec in enumerate(ECOLS):
        write_cell(ws3, r, 4+k, round(row[ec],1), bg=bg)
    chg = row["EBITDA_FY24"]-row["EBITDA_FY19"]
    chg_bg = "D5F5E3" if chg>=0 else "FADBD8"
    write_cell(ws3, r, 10, round(chg,1), bold=True, bg=chg_bg)

for col_letter in ["D","E","F","G","H","I"]:
    ws3.conditional_formatting.add(
        f"{col_letter}3:{col_letter}{2+len(df)}",
        ColorScaleRule(start_type="min", start_color="FFFFFF",
                       end_type="max",   end_color="1E8449"))

col_w(ws3, [9,22,16,8,8,8,8,8,8,14])

# ---------------------------------------------------------------
# SHEET 4: Revenue Data
# ---------------------------------------------------------------
ws4 = wb.create_sheet("Revenue Data")
ws4.sheet_view.showGridLines = False

ws4.merge_cells("A1:J1")
ws4["A1"].value = "Revenue by Company and Year (INR Crores)"
ws4["A1"].font = Font(bold=True, size=12, color="FFFFFF")
ws4["A1"].fill = PatternFill("solid", fgColor="2C3E50")
ws4["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws4.row_dimensions[1].height = 26

h4 = ["Ticker","Company","Sub-Sector",
      "FY19","FY20","FY21","FY22","FY23","FY24","5Y CAGR (%)"]
for j,h in enumerate(h4,1):
    write_cell(ws4, 2, j, h, bold=True, sz=9, bg="ECF0F1", wrap=True)
ws4.row_dimensions[2].height = 24

dfr = df.sort_values("Rev_CAGR_5Y", ascending=False).reset_index(drop=True)
for i, row in dfr.iterrows():
    r = i+3
    bg = "F8F9F9" if i%2==0 else "FFFFFF"
    write_cell(ws4, r, 1, row["Ticker"], bg=bg)
    write_cell(ws4, r, 2, row["Company"], align="left", bg=bg)
    write_cell(ws4, r, 3, row["Sub_Sector"], align="left", bg=bg)
    for k,rc in enumerate(RCOLS):
        write_cell(ws4, r, 4+k, int(row[rc]), bg=bg, num_fmt="#,##0")
    cagr_bg = ("D5F5E3" if row["Rev_CAGR_5Y"]>15 else
               "FEF9E7" if row["Rev_CAGR_5Y"]>8 else "FADBD8")
    write_cell(ws4, r, 10, round(row["Rev_CAGR_5Y"],1),
               bold=True, bg=cagr_bg)

ws4.conditional_formatting.add(
    f"D3:I{2+len(df)}",
    ColorScaleRule(start_type="min", start_color="FFFFFF",
                   end_type="max",   end_color="1A5276"))
col_w(ws4, [9,22,16,10,10,10,10,10,10,12])

# ---------------------------------------------------------------
# SHEET 5: Cluster Results
# ---------------------------------------------------------------
ws5 = wb.create_sheet("Cluster Results")
ws5.sheet_view.showGridLines = False

ws5.merge_cells("A1:I1")
ws5["A1"].value = "K-Means Clustering Results (k=4) - Company Segmentation"
ws5["A1"].font = Font(bold=True, size=12, color="FFFFFF")
ws5["A1"].fill = PatternFill("solid", fgColor="6C3483")
ws5["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws5.row_dimensions[1].height = 26

ws5.merge_cells("A2:I2")
ws5["A2"].value = ("Features: Revenue CAGR, Avg EBITDA, Avg PAT, Avg ROE, "
                   "Export %, R&D %, Capex % | Standardized (Z-score) before fitting | k chosen via elbow method")
ws5["A2"].font = Font(size=8, italic=True, color="555555")
ws5["A2"].alignment = Alignment(horizontal="center")
ws5.row_dimensions[2].height = 16

h5 = ["Cluster","Profile","Ticker","Company","Sub-Sector",
      "Avg EBITDA (%)","Avg ROE (%)","Rev CAGR (%)","Composite Score"]
for j,h in enumerate(h5,1):
    write_cell(ws5, 3, j, h, bold=True, sz=9, bg="ECF0F1", wrap=True)
ws5.row_dimensions[3].height = 28

# light tints per cluster, nothing loud
ctints = {1:"EAF4FB",2:"EAFAF1",3:"FEF9E7",4:"FDEDEC"}
dfc = df.sort_values(["Cluster","Avg_EBITDA"], ascending=[True,False])
r = 4
for cid in sorted(dfc["Cluster"].unique()):
    sub = dfc[dfc["Cluster"]==cid]
    bg  = ctints[cid]
    for _, row in sub.iterrows():
        write_cell(ws5, r, 1, int(cid), bg=bg)
        write_cell(ws5, r, 2, cnames[cid], align="left", bg=bg)
        write_cell(ws5, r, 3, row["Ticker"], bg=bg)
        write_cell(ws5, r, 4, row["Company"], align="left", bg=bg)
        write_cell(ws5, r, 5, row["Sub_Sector"], align="left", bg=bg)
        write_cell(ws5, r, 6, round(row["Avg_EBITDA"],1), bg=bg)
        write_cell(ws5, r, 7, round(row["Avg_ROE"],1), bg=bg)
        write_cell(ws5, r, 8, round(row["Rev_CAGR_5Y"],1), bg=bg)
        write_cell(ws5, r, 9, round(row["Score"],1), bold=True, bg=bg)
        r += 1
    r += 1  # blank row between clusters

col_w(ws5, [8,24,9,22,16,12,10,12,14])

wb.save(OUT)
print(f"Excel saved: {OUT}")
print("Sheets: Master Data | Sub-sector Summary | EBITDA Trends | Revenue Data | Cluster Results")
